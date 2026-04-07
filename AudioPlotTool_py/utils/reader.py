from __future__ import annotations
# -*- coding: utf-8 -*-
"""
reader.py  –  从 AP 导出的 Excel sheet 读取频率轴和 sweep 数据

Excel 格式：
  第1行（表头）: TestDate | TestTime | project | Hz | Result | 20 | 21.2 | ...
  第2行起:       日期     | 时间     | project | 单位 | Pass  | 数值 ...
"""

import numpy as np
import openpyxl
from typing import Optional

# ── Sheet 名称定义 ────────────────────────────────────────
SHEET_NAMES = {
    "FR_L":          "Freqresp -ear L",
    "FR_R":          "Freqresp -ear R",
    "THD_L":         "THD 2-5 -ear L",
    "THD_R":         "THD 2-5 -ear R",
    "RB_L":          "Rub&Buzz Harmonic 10-35 -ear L",
    "RB_R":          "Rub&Buzz Harmonic 10-35 -ear R",
    "Leakage_mic2":  "Freqresp -mic 2",
}


def get_sheet_names(file_path: str) -> list[str]:
    """返回 Excel 文件中所有 sheet 名称。"""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    except Exception as e:
        print(f"  [警告] 无法读取 sheet 列表：{file_path}  ({e})")
        return []


def read_ap_sheet(file_path: str, sheet_name: str
                  ) -> tuple[Optional[np.ndarray], Optional[np.ndarray]]:
    """
    读取指定 sheet，返回 (freq, data)。
      freq : shape (N,)   频率轴 Hz
      data : shape (M, N) M 行 sweep 数据
    若 sheet 不存在或读取失败，返回 (None, None)。
    """
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  [警告] 无法打开文件：{file_path}  ({e})")
        return None, None

    try:
        if sheet_name not in wb.sheetnames:
            return None, None

        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if len(rows) < 2:
        return None, None

    header = rows[0]

    # 定位数据起始列：找 'Hz' 标签，数据列 = Hz列 + 2
    data_col: Optional[int] = None
    for c, val in enumerate(header):
        if isinstance(val, str) and val.strip().upper() == "HZ":
            data_col = c + 2
            break

    if data_col is None:
        print(f"  [警告] sheet '{sheet_name}' 中未找到 'Hz' 列标签，"
              f"无法确定数据起始列，跳过。")
        return None, None

    # 频率轴：第1行从 data_col 起（跳过 None，兼容合并单元格）
    freq_vals = []
    for v in header[data_col:]:
        if isinstance(v, (int, float)) and v is not None:
            freq_vals.append(float(v))
        # None 或非数值均跳过（合并单元格可能产生 None 间隙）

    if not freq_vals:
        return None, None

    freq = np.array(freq_vals, dtype=float)
    n    = len(freq)

    # sweep 数据：第2行起
    data_rows = []
    for row in rows[1:]:
        vals = row[data_col: data_col + n]
        row_data = []
        for v in vals:
            if isinstance(v, (int, float)) and v is not None:
                row_data.append(float(v))
            else:
                row_data.append(np.nan)
        # 去除全 NaN 行
        arr = np.array(row_data, dtype=float)
        if not np.all(np.isnan(arr)):
            data_rows.append(arr)

    if not data_rows:
        return None, None

    data = np.vstack(data_rows)   # shape (M, N)
    return freq, data


def read_channel(file_path: str, sheet_l: str, sheet_r: str
                 ) -> tuple[Optional[np.ndarray], Optional[np.ndarray], str]:
    """
    优先读 L 声道，失败则读 R 声道。
    返回 (freq, data, channel_label)。
    """
    freq, data = read_ap_sheet(file_path, sheet_l)
    if freq is not None:
        return freq, data, "L"
    freq, data = read_ap_sheet(file_path, sheet_r)
    if freq is not None:
        return freq, data, "R"
    return None, None, ""
