#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ar_glasses_fr_plot.py
输入 AP 导出的 Excel，取 1kHz 处最接近 75 dBSPL 的行，
绘制 AR 眼镜扬声器频响示意图（H_glass / H_target / ΔH 双面板）。

依赖：pip3 install matplotlib numpy openpyxl

用法：
    python3 ar_glasses_fr_plot.py [excel_path]
"""

import os
import sys
import platform
import numpy as np

import matplotlib
try:
    matplotlib.use("TkAgg")
    import tkinter as _tk; del _tk
except Exception:
    matplotlib.use("Agg")

import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import matplotlib.ticker as ticker
import matplotlib.font_manager as fm
import openpyxl
from tkinter import Tk, filedialog

X_LIM = (50, 20000)
FC_LOW, FC_HIGH = 300, 800

# ── 哈曼目标曲线（内嵌，无需外部文件）──────────────────────────
_HARMAN_F = np.array([
    20,21,22,24,25,27,28,30,32,34,36,38,40,43,45,48,50,53,56,60,
    63,67,71,75,80,85,90,95,100,106,112,118,125,132,140,150,160,170,
    180,190,200,212,224,236,250,265,280,300,315,335,355,375,400,425,
    450,475,500,530,560,600,630,670,710,750,800,850,900,950,1000,
    1060,1120,1180,1250,1320,1400,1500,1600,1700,1800,1900,2000,
    2120,2240,2360,2500,2650,2800,3000,3150,3350,3550,3750,4000,
    4250,4500,4750,5000,5300,5600,6000,6300,6700,7100,7500,8000,
    8500,9000,9500,10000,10600,11200,11800,12500,13200,14000,15000,
    16000,17000,18000,19000,20000
], dtype=float)

_HARMAN_SPL = np.array([
    79.198164,79.366553,79.487179,79.763854,79.95338,79.95338,
    79.95338,79.95338,79.894517,79.804802,79.72028,79.684305,
    79.605611,79.48757,79.452838,79.340182,79.265079,79.14489,
    79.001393,78.839659,78.632255,78.470226,78.321678,78.132349,
    77.88509,77.622378,77.404401,76.981999,76.689977,76.336708,
    75.96253,75.524476,75.107415,74.825175,74.410302,74.070032,
    73.659674,73.373205,72.970608,72.727273,72.397979,72.211561,
    72.027972,72.027972,72.027972,72.027972,72.027972,72.027972,
    72.027972,72.027972,72.027972,72.027972,72.027972,72.027972,
    72.027972,72.027972,72.027972,72.031952,72.261072,72.35788,
    72.497367,72.494172,72.727273,72.703714,72.672175,72.640636,
    72.651224,72.727273,72.727273,72.890112,73.18639,73.344657,
    73.659674,73.886391,74.358974,74.754214,75.345012,76.05981,
    76.689977,77.517831,78.243872,78.718778,79.382006,79.95338,
    80.41958,80.652681,80.885781,81.21005,81.546311,81.585082,
    81.585082,81.585082,81.234489,80.892607,80.41958,79.791324,
    79.361718,78.822961,78.509795,78.102294,77.646404,77.156177,
    76.665234,76.005857,75.524476,74.816761,74.008044,73.359139,
    72.516254,71.574738,70.445737,69.25067,68.075318,66.954604,
    66.005028,65.501166,64.930836,63.80885,61.758311,58.047584,
    54.252077
], dtype=float)


# ── 1. 中文字体 ─────────────────────────────────────────────────
def _setup_font():
    sys_name = platform.system()
    if sys_name == "Darwin":
        cands = ["PingFang SC", "Heiti SC", "STHeiti", "Arial Unicode MS"]
    elif sys_name == "Windows":
        cands = ["Microsoft YaHei", "SimHei"]
    else:
        cands = ["WenQuanYi Micro Hei", "Noto Sans CJK SC", "DejaVu Sans"]
    available = {f.name for f in fm.fontManager.ttflist}
    for name in cands:
        if name in available:
            plt.rcParams["font.family"] = name
            break
    plt.rcParams["axes.unicode_minus"] = False
    plt.rcParams["mathtext.fontset"] = "cm"


# ── 2. 读取 Excel FR 数据 ────────────────────────────────────────
def read_fr_sheet(path, sheet="Freqresp -ear L"):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet}' not found in {path}")
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = rows[0]
    data_col = None
    for c, v in enumerate(header):
        if isinstance(v, str) and v.strip().upper() == "HZ":
            data_col = c + 2
            break
    if data_col is None:
        raise ValueError("Cannot find 'Hz' column in header row")

    freq = np.array([float(v) for v in header[data_col:]
                     if isinstance(v, (int, float)) and v is not None])
    n = len(freq)

    data_rows = []
    for row in rows[1:]:
        arr = np.array([
            float(v) if isinstance(v, (int, float)) and v is not None else np.nan
            for v in row[data_col: data_col + n]
        ])
        if not np.all(np.isnan(arr)):
            data_rows.append(arr)

    return freq, np.vstack(data_rows)


# ── 3. 选行：1kHz 处最接近 75 dBSPL ─────────────────────────────
def select_row_near_75(freq, data, target_freq=1000.0, target_spl=75.0):
    idx_f = int(np.argmin(np.abs(freq - target_freq)))
    spl_at_f = data[:, idx_f]
    best = int(np.argmin(np.abs(spl_at_f - target_spl)))
    return best, freq[idx_f], float(spl_at_f[best])


# ── 4. 对齐哈曼曲线（500~2000Hz 均值对齐）──────────────────────
def align_harman(ref_freq, ref_data, align_low=500.0, align_high=2000.0):
    h_on_ref = np.interp(ref_freq, _HARMAN_F, _HARMAN_SPL)
    mask = (ref_freq >= align_low) & (ref_freq <= align_high)
    if np.any(mask):
        offset = np.nanmean(ref_data[mask]) - np.nanmean(h_on_ref[mask])
    else:
        offset = 0.0
    return h_on_ref + offset


# ── 5. X轴工具 ──────────────────────────────────────────────────
def _fmt_hz(val, _pos=None):
    return f"{val/1000:g}k" if val >= 1000 else f"{val:g}"


def _apply_log_xaxis(ax, xlabel=True):
    ax.set_xscale("log")
    ax.set_xlim(X_LIM)
    ticks = [50, 100, 200, 500, 1000, 2000, 5000, 10000, 20000]
    ax.set_xticks(ticks)
    ax.xaxis.set_major_formatter(ticker.FuncFormatter(_fmt_hz))
    ax.xaxis.set_minor_formatter(ticker.NullFormatter())
    if xlabel:
        ax.set_xlabel("Frequency (Hz)", fontsize=12)
    ax.grid(True, which="major", alpha=0.25, linewidth=0.7)
    ax.grid(True, which="minor", alpha=0.10, linewidth=0.4)
    for sp in ax.spines.values():
        sp.set_linewidth(0.8)
    ax.tick_params(labelsize=10)


# ── 6a. 概览图：H_glass + H_target ──────────────────────────────
def plot_overview(freq, h_glass, h_target, spl_1k, out_path):
    mask = (freq >= X_LIM[0]) & (freq <= X_LIM[1])
    xf, yg, yt = freq[mask], h_glass[mask], h_target[mask]

    y_min = 20.0
    y_max = float(np.ceil((np.nanmax(yg) + 12) / 10) * 10)

    fig, ax = plt.subplots(figsize=(13, 6.5), facecolor="white")
    fig.patch.set_facecolor("white")
    fig.suptitle(
        "AR眼镜扬声器频响曲线 $H_{glass}(f)$ 与哈曼目标曲线 $H_{target}(f)$",
        fontsize=13, fontweight="bold", y=0.98,
    )

    ax.set_facecolor("white")
    ax.set_ylim(y_min, y_max)
    ax.axvspan(FC_LOW, FC_HIGH, alpha=0.14, color="orange", zorder=0)
    ax.plot(xf, yg, color="#1f77b4", linewidth=2.0,
            label=f"$H_{{glass}}(f)$  AR眼镜扬声器频响曲线（整数标识）")
    ax.plot(xf, yt, color="black", linewidth=1.6, linestyle="--",
            label="$H_{{target}}(f)$  哈曼目标曲线")
    ax.text(
        np.sqrt(FC_LOW * FC_HIGH),
        y_min + (y_max - y_min) * 0.88,
        "分频点范围\n300~800 Hz",
        fontsize=9, color="#7b3800", ha="center", va="top",
        bbox=dict(boxstyle="round,pad=0.3", fc="lightyellow",
                  ec="orange", alpha=0.88),
        zorder=5,
    )
    ax.set_ylabel("SPL (dB)", fontsize=12)
    ax.legend(loc="lower left", fontsize=9, frameon=True)
    _apply_log_xaxis(ax, xlabel=True)

    fig.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches="tight", facecolor="white")
    print(f"[OK] 概览图已保存：{out_path}")


# ── 6b. 调音参考图：ΔH(f)，200Hz~20kHz，1dB 刻度 ────────────────
def plot_tuning(freq, h_glass, h_target, spl_1k, out_path):
    X_TUNING = (200, 20000)
    mask = (freq >= X_TUNING[0]) & (freq <= X_TUNING[1])
    xf = freq[mask]
    yd = (h_glass - h_target)[mask]

    # Y 轴范围：贴合数据，边界对齐到 5dB 整数倍
    valid = yd[np.isfinite(yd)]
    y_low  = int(np.floor(np.nanmin(valid) - 2) // 5) * 5
    y_high = int(np.ceil( np.nanmax(valid) + 2) // 5 + 1) * 5

    fig, ax = plt.subplots(figsize=(14, 8), facecolor="white")
    fig.patch.set_facecolor("white")
    fig.suptitle(
        "调音参考  $\\Delta H(f) = H_{glass}(f) - H_{target}(f)$"
        f"    （1kHz ≈ {spl_1k:.0f} dBSPL）",
        fontsize=13, fontweight="bold", y=0.98,
    )

    ax.set_facecolor("#fafafa")

    # 分频点阴影
    ax.axvspan(FC_LOW, FC_HIGH, alpha=0.12, color="orange", zorder=0)

    # ±填色：直观区分需要加/减的区域
    ax.fill_between(xf, yd, 0, where=(yd >= 0),
                    color="#2ca02c", alpha=0.18, zorder=1)
    ax.fill_between(xf, yd, 0, where=(yd <= 0),
                    color="#d62728", alpha=0.18, zorder=1)

    # 零线（加粗、实线，调音基准线）
    ax.axhline(0, color="black", linewidth=1.5, zorder=3)

    # ΔH 曲线
    ax.plot(xf, yd, color="#1a1a1a", linewidth=1.8, zorder=4,
            label="$\\Delta H(f)$（正值 = 实测高于目标，需衰减；负值 = 低于目标，需提升）")

    # ── Y 轴：1dB 精度 ──────────────────────────────────────────
    ax.set_ylim(y_low, y_high)
    ax.yaxis.set_major_locator(ticker.MultipleLocator(5))
    ax.yaxis.set_minor_locator(ticker.MultipleLocator(1))
    ax.yaxis.set_major_formatter(ticker.FuncFormatter(
        lambda v, _: f"{int(v):+d} dB" if v != 0 else "0 dB"))
    ax.set_ylabel("$\\Delta H$ (dB)", fontsize=12)

    # 网格：主刻度稍重，1dB 副刻度浅
    ax.grid(True, which="major", color="gray", alpha=0.35, linewidth=0.8)
    ax.grid(True, which="minor", color="gray", alpha=0.12, linewidth=0.4)

    # ── X 轴 ─────────────────────────────────────────────────────
    ax.set_xscale("log")
    ax.set_xlim(X_TUNING)
    x_ticks = [200, 315, 500, 800, 1000, 2000, 3150, 5000, 8000, 10000, 20000]
    ax.set_xticks(x_ticks)
    ax.xaxis.set_major_formatter(ticker.FuncFormatter(_fmt_hz))
    ax.xaxis.set_minor_formatter(ticker.NullFormatter())
    ax.set_xlabel("Frequency (Hz)", fontsize=12)

    # 右侧注释标签
    kw = dict(xycoords="axes fraction", textcoords="axes fraction",
              fontsize=9, va="center", ha="left",
              annotation_clip=False,
              arrowprops=dict(arrowstyle="->", lw=1.1))
    ax.annotate("高于目标（建议衰减）",
                xy=(0.99, 0.72), xytext=(1.005, 0.72),
                color="#2ca02c", **kw)
    ax.annotate("低于目标（建议提升）",
                xy=(0.99, 0.28), xytext=(1.005, 0.28),
                color="#d62728", **kw)

    for sp in ax.spines.values():
        sp.set_linewidth(0.8)
    ax.tick_params(which="major", labelsize=10, length=5)
    ax.tick_params(which="minor", labelsize=0,  length=3)

    ax.legend(loc="upper right", fontsize=8.5, frameon=True,
              framealpha=0.9)

    fig.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches="tight", facecolor="white")
    print(f"[OK] 调音参考图已保存：{out_path}")


# ── 主入口 ──────────────────────────────────────────────────────
def main():
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    else:
        root = Tk()
        root.withdraw()
        root.wm_attributes("-topmost", True)
        excel_path = filedialog.askopenfilename(
            title="选择 AP 导出的 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xlsm"), ("所有文件", "*.*")],
        )
        root.destroy()

    if not excel_path:
        print("未选择文件，程序退出。")
        sys.exit(0)

    if not os.path.isfile(excel_path):
        print(f"[错误] 找不到文件：{excel_path}")
        sys.exit(1)

    print(f"读取：{excel_path}")
    freq, data = read_fr_sheet(excel_path)
    print(f"  频率轴 {freq[0]:.0f}~{freq[-1]:.0f} Hz，{len(freq)} 点，{data.shape[0]} 行 sweep")

    row_idx, f_1k, spl_1k = select_row_near_75(freq, data)
    print(f"  选取第 {row_idx} 行：{f_1k:.0f} Hz 处 SPL = {spl_1k:.1f} dB（最接近 75 dBSPL）")

    h_glass  = data[row_idx]
    h_target = align_harman(freq, h_glass)

    base_dir = os.path.expanduser("/Users/sly/Desktop/Audio Test Export")
    os.makedirs(base_dir, exist_ok=True)
    stem     = os.path.splitext(os.path.basename(excel_path))[0]

    _setup_font()
    plot_overview(freq, h_glass, h_target, spl_1k,
                  os.path.join(base_dir, f"{stem}_overview.png"))
    plot_tuning(freq, h_glass, h_target, spl_1k,
                os.path.join(base_dir, f"{stem}_tuning.png"))
    plt.show()


if __name__ == "__main__":
    main()
